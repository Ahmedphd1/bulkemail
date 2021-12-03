import httplib2
import os
import base64
from email import encoders
#needed for attachment
import smtplib
import mimetypes
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
import ssl
import configparser

#List of all mimetype per extension: http://help.dottoro.com/lapuadlp.php  or http://mime.ritey.com/

from apiclient import errors, discovery  #needed for gmail service

config = configparser.ConfigParser()
config.read('config.ini')

excelpath = str(config['variables']['excelpath'])
ditnavn = str(config['variables']['navn'])
email = str(config['variables']['email'])
ditpassword = str(config['variables']['password'])

def getemaillist():
    from openpyxl import load_workbook
    wb = load_workbook(excelpath)  # Work Book
    ws = wb.get_sheet_by_name('Sheet0')  # Work Sheet
    column = ws['H']  # Column
    column_list = [column[x].value for x in range(len(column))]

    mylist = [column_list[x] for x in range(len(column_list)) if column_list[x] != None]

    mylist.pop(0)

    return mylist

def createmessage(files):
    message = MIMEMultipart("alternative")

    for attached_files in files:

        my_mimetype, encoding = mimetypes.guess_type(attached_files)

        if my_mimetype is None or encoding is not None:
            my_mimetype = 'application/octet-stream'

        main_type, sub_type = my_mimetype.split('/', 1)

        if main_type == 'text':
            print("text")
            temp = open(attached_files, 'r')  # 'rb' will send this error: 'bytes' object has no attribute 'encode'
            attachment = MIMEText(temp.read(), _subtype=sub_type)
            temp.close()

        elif main_type == 'image':
            print("image")
            temp = open(attached_files, 'rb')
            attachment = MIMEImage(temp.read(), _subtype=sub_type)
            temp.close()

        elif main_type == 'audio':
            print("audio")
            temp = open(attached_files, 'rb')
            attachment = MIMEAudio(temp.read(), _subtype=sub_type)
            temp.close()

        elif main_type == 'application' and sub_type == 'pdf':
            temp = open(attached_files, 'rb')
            attachment = MIMEApplication(temp.read(), _subtype=sub_type)
            temp.close()

        else:
            attachment = MIMEBase(main_type, sub_type)
            temp = open(attached_files, 'rb')
            attachment.set_payload(temp.read())
            temp.close()

        # -----3.3 encode the attachment, add a header and attach it to the message
        # encoders.encode_base64(attachment)  #not needed (cf. randomfigure comment)
        # https://docs.python.org/3/library/email-examples.html

        filename = os.path.basename(attached_files)
        attachment.add_header('Content-Disposition', 'attachment', filename=filename)  # name preview in email
        message.attach(attachment)
    return message

def sendmessage(targetemail, message):
    global navn
    global email
    global password

    navn = ditnavn

    text = f"\n Hej {targetemail.split('@')[1]} \n \n Jeg vil gerne søge praktikplads hos jer \n \n Jeg har dokumenteret mit CV og Ansøgning i mailen. Jeg er elev på tec og studere datateknikker med speciale i programmering \n \n Venlig hilsen {navn}"

    message.attach(MIMEText(text, 'plain'))

    sender_email = email
    # receiver_email = "receiver_mail@gmail.com"
    password = ditpassword

    message["Subject"] = "Vedrørende praktikplads"
    message["From"] = sender_email

    context = ssl.create_default_context()
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(sender_email, password)
        server.sendmail(sender_email, targetemail, message.as_string())


